import json
import smtplib
import socket
from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.exc import SQLAlchemyError
from typer.testing import CliRunner

from app.cli.main import app
from app.core.database import SessionLocal
from app.modules.company.models import Company
from app.modules.company_enrichment.models import CompanyEnrichment
from app.modules.contact.models import Contact
from app.modules.crm.excel_export import CRMExcelExportError, CRMExcelExportService
from app.modules.crm.outreach_batch import OutreachBatchStatus, OutreachBatchWorkflow
from app.modules.email_delivery.manual_models import ManualEmailSendRecord
from app.modules.email_delivery.manual_repository import ManualEmailSendRecordRepository
from app.modules.email_delivery.manual_schemas import (
    ConfirmedExternalManualEmailSendBatchCommand,
)
from app.modules.email_delivery.manual_service import (
    ManualOutreachPersistenceError,
    ManualOutreachService,
)
from app.modules.email_delivery.models import EmailDeliveryAttempt
from app.modules.email_draft.context import build_content_hash
from app.modules.email_draft.models import EmailDraft
from app.modules.lead.models import Lead
from app.modules.project.models import Project
from app.modules.task.models import Task

NOW = datetime(2026, 8, 25, 8, 0, tzinfo=UTC)
runner = CliRunner()


def _batch(count: int, *, person_position: int | None = None) -> dict[str, object]:
    with SessionLocal() as session:
        project = Project(name="Batch export")
        session.add(project)
        session.flush()
        drafts: list[int] = []
        leads: list[int] = []
        tasks: list[int] = []
        contacts: list[int] = []
        for position in range(1, count + 1):
            recipient = f"recipient-{position}@example.test"
            company = Company(project_id=project.id, name=f"Studio {position}")
            session.add(company)
            session.flush()
            session.add(
                CompanyEnrichment(
                    company_id=company.id,
                    enrichment_status="SUCCEEDED",
                    email=recipient,
                    source_url=f"https://studio-{position}.example.test/contact",
                )
            )
            contact = None
            if position == person_position:
                contact = Contact(
                    company_id=company.id,
                    first_name="Person",
                    email=recipient,
                    source="MANUAL_VERIFIED",
                )
                session.add(contact)
                session.flush()
                contacts.append(contact.id)
            lead = Lead(
                company_id=company.id,
                contact_id=contact.id if contact else None,
                status="NEW",
                source="MANUAL" if contact else "COMPANY_SCOPED_OUTREACH",
            )
            session.add(lead)
            session.flush()
            task = Task(lead_id=lead.id, title="Prepare outreach", status="TODO")
            session.add(task)
            session.flush()
            subject = f"Subject {position}"
            body = f"Body {position}"
            prompt_version = "test-v1"
            draft = EmailDraft(
                project_id=project.id,
                company_id=company.id,
                contact_id=contact.id if contact else None,
                lead_id=lead.id,
                task_id=task.id,
                recipient_email=recipient,
                recipient_name="Person" if contact else f"Studio {position} team",
                recipient_role=None,
                sender_name="Alex",
                sender_company="Bohemia Bali",
                generation_tone="warm",
                generation_purpose="Outreach",
                generation_value_proposition=None,
                subject=subject,
                text_body=body,
                language="en",
                prompt_version=prompt_version,
                provider="fake",
                model="fake",
                context_fingerprint=f"{position:064x}",
                request_fingerprint=f"{position + 100:064x}",
                content_hash=build_content_hash(
                    recipient_email=recipient,
                    subject=subject,
                    text_body=body,
                    prompt_version=prompt_version,
                ),
                status="DRAFT",
            )
            session.add(draft)
            session.flush()
            drafts.append(draft.id)
            leads.append(lead.id)
            tasks.append(task.id)
        session.commit()
        return {
            "project": project.id,
            "drafts": tuple(drafts),
            "leads": tuple(leads),
            "tasks": tuple(tasks),
            "contacts": tuple(contacts),
        }


def _invoke(ids: dict[str, object], destination: Path, *, confirm: bool = True):
    args = ["crm", "export-outreach-batch", "--project-id", str(ids["project"])]
    for draft_id in ids["drafts"]:
        args.extend(("--email-draft-id", str(draft_id)))
    args.extend(("--output-file", str(destination), "--output", "json"))
    if confirm:
        args.append("--confirm")
    return runner.invoke(app, args)


def _draft_snapshot(draft: EmailDraft) -> tuple[object, ...]:
    return (
        draft.subject,
        draft.text_body,
        draft.recipient_email,
        draft.contact_id,
        "PERSON" if draft.contact_id is not None else "COMPANY",
        draft.content_hash,
        draft.context_fingerprint,
        draft.request_fingerprint,
        draft.provider,
        draft.model,
        draft.prompt_version,
        draft.status,
    )


def test_batch_without_confirmation_is_preview_only(tmp_path: Path) -> None:
    ids = _batch(1)
    destination = tmp_path / "preview.xlsx"
    result = _invoke(ids, destination, confirm=False)
    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload["status"] == "CONFIRMATION_REQUIRED"
    assert payload["email_draft_ids"] == list(ids["drafts"])
    assert not destination.exists()
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0
        draft = session.get(EmailDraft, ids["drafts"][0])
        assert draft is not None and draft.delivery_mode is None


@pytest.mark.parametrize("count", [1, 5])
def test_confirmed_batch_records_all_once_and_exports_sent(tmp_path: Path, count: int) -> None:
    ids = _batch(count, person_position=1)
    with SessionLocal() as session:
        snapshots = {
            draft_id: _draft_snapshot(session.get_one(EmailDraft, draft_id))
            for draft_id in ids["drafts"]
        }
    destination = tmp_path / f"batch-{count}.xlsx"
    result = _invoke(ids, destination)
    assert result.exit_code == 0, result.output
    payload = json.loads(result.stdout)
    assert payload["status"] == "COMPLETE"
    assert len(payload["items"]) == count
    assert {item["outreach_status"] for item in payload["items"]} == {"MANUALLY_SENT"}
    assert {item["recipient_type"] for item in payload["items"]} >= {"PERSON"}
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == count
        assert session.scalar(select(func.count()).select_from(EmailDeliveryAttempt)) == 0
        for draft_id, lead_id, task_id in zip(
            ids["drafts"], ids["leads"], ids["tasks"], strict=True
        ):
            draft = session.get(EmailDraft, draft_id)
            lead = session.get(Lead, lead_id)
            task = session.get(Task, task_id)
            assert draft is not None and draft.status == "DRAFT"
            assert draft.delivery_mode == "MANUAL"
            assert _draft_snapshot(draft) == snapshots[draft_id]
            assert lead is not None and lead.status == "NEW"
            assert task is not None and task.status == "TODO"
    workbook = load_workbook(destination, read_only=True)
    headers = {cell.value: index for index, cell in enumerate(workbook["Outreach"][1])}
    rows = list(workbook["Outreach"].iter_rows(min_row=2, values_only=True))
    assert len(rows) == count
    assert {row[headers["Outreach Status"]] for row in rows} == {"MANUALLY_SENT"}
    assert all(row[headers["Manual Sent At"]] is not None for row in rows)
    workbook.close()


def test_destination_preflight_failure_does_not_record_batch(tmp_path: Path) -> None:
    ids = _batch(1)
    blocked_parent = tmp_path / "not-a-directory"
    blocked_parent.write_text("block", encoding="utf-8")
    destination = blocked_parent / "batch.xlsx"
    result = _invoke(ids, destination)
    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["status"] == "EXPORT_FAILED_BEFORE_MUTATION"
    assert not destination.exists()
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0
        draft = session.get_one(EmailDraft, ids["drafts"][0])
        assert draft.delivery_mode is None


def test_text_preview_lists_every_planned_manual_send(tmp_path: Path) -> None:
    ids = _batch(2, person_position=1)
    args = ["crm", "export-outreach-batch", "--project-id", str(ids["project"])]
    for draft_id in ids["drafts"]:
        args.extend(("--email-draft-id", str(draft_id)))
    args.extend(("--output-file", str(tmp_path / "preview.xlsx")))
    result = runner.invoke(app, args)
    assert result.exit_code == 0, result.output
    assert "Status: CONFIRMATION_REQUIRED" in result.stdout
    assert f"Project ID: {ids['project']}" in result.stdout
    assert "Batch size: 2" in result.stdout
    assert result.stdout.count("Action: RECORD_MANUAL_SEND") == 2
    assert "Recipient type: PERSON" in result.stdout
    assert "Recipient type: COMPANY" in result.stdout
    assert "Current outreach: READY_FOR_MANUAL_SEND" in result.stdout
    for draft_id in ids["drafts"]:
        assert f"Draft {draft_id}" in result.stdout
    assert "recipient-1@example.test" in result.stdout
    assert "recipient-2@example.test" in result.stdout


def test_non_master_selected_draft_exports_complete_relationships(tmp_path: Path) -> None:
    ids = _batch(1)
    first_draft_id = ids["drafts"][0]
    with SessionLocal() as session:
        first = session.get_one(EmailDraft, first_draft_id)
        second_lead = Lead(
            company_id=first.company_id,
            contact_id=None,
            status="NEW",
            source="COMPANY_SCOPED_OUTREACH",
        )
        session.add(second_lead)
        session.flush()
        second_task = Task(lead_id=second_lead.id, title="Newer outreach", status="TODO")
        session.add(second_task)
        session.flush()
        second_subject = "Newer subject"
        second_body = "Newer body"
        second = EmailDraft(
            project_id=first.project_id,
            company_id=first.company_id,
            contact_id=None,
            lead_id=second_lead.id,
            task_id=second_task.id,
            recipient_email=first.recipient_email,
            recipient_name=first.recipient_name,
            recipient_role=None,
            sender_name=first.sender_name,
            sender_company=first.sender_company,
            generation_tone=first.generation_tone,
            generation_purpose=first.generation_purpose,
            generation_value_proposition=None,
            subject=second_subject,
            text_body=second_body,
            language=first.language,
            prompt_version=first.prompt_version,
            provider=first.provider,
            model=first.model,
            context_fingerprint="e" * 64,
            request_fingerprint="f" * 64,
            content_hash=build_content_hash(
                recipient_email=first.recipient_email,
                subject=second_subject,
                text_body=second_body,
                prompt_version=first.prompt_version,
            ),
            status="DRAFT",
        )
        session.add(second)
        session.commit()
        assert second.id > first_draft_id

    destination = tmp_path / "selected-older.xlsx"
    result = _invoke(ids, destination)
    assert result.exit_code == 0, result.output
    workbook = load_workbook(destination, read_only=True)
    try:
        expected = {
            "Sales Leads": ("Draft ID", first_draft_id),
            "Companies": ("Company ID", session_company_id(ids)),
            "Tasks": ("Task ID", ids["tasks"][0]),
            "Outreach": ("Draft ID", first_draft_id),
        }
        for sheet_name, (header, value) in expected.items():
            sheet = workbook[sheet_name]
            headers = {cell.value: index for index, cell in enumerate(sheet[1])}
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            assert [row[headers[header]] for row in rows] == [value]
        assert workbook["Contacts"].max_row == 1
    finally:
        workbook.close()


def session_company_id(ids: dict[str, object]) -> int:
    with SessionLocal() as session:
        draft = session.get_one(EmailDraft, ids["drafts"][0])
        return draft.company_id


def test_person_and_company_scope_are_preserved_exactly(tmp_path: Path) -> None:
    ids = _batch(2, person_position=1)
    with SessionLocal() as session:
        before = {
            draft_id: _draft_snapshot(session.get_one(EmailDraft, draft_id))
            for draft_id in ids["drafts"]
        }
    result = _invoke(ids, tmp_path / "scopes.xlsx")
    assert result.exit_code == 0, result.output
    with SessionLocal() as session:
        after = {
            draft_id: _draft_snapshot(session.get_one(EmailDraft, draft_id))
            for draft_id in ids["drafts"]
        }
        assert after == before
        person = session.get_one(EmailDraft, ids["drafts"][0])
        company = session.get_one(EmailDraft, ids["drafts"][1])
        assert person.contact_id == ids["contacts"][0]
        assert company.contact_id is None
        assert person.delivery_mode == company.delivery_mode == "MANUAL"


def test_ordinary_crm_export_remains_read_only(tmp_path: Path) -> None:
    ids = _batch(1)
    result = runner.invoke(
        app,
        [
            "crm",
            "export-excel",
            "--project-id",
            str(ids["project"]),
            "--output-file",
            str(tmp_path / "ordinary.xlsx"),
        ],
    )
    assert result.exit_code == 0, result.output
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0
        assert session.get_one(EmailDraft, ids["drafts"][0]).delivery_mode is None


def test_batch_conflict_rejects_every_draft_before_mutation(tmp_path: Path) -> None:
    ids = _batch(2)
    first, second = ids["drafts"]
    with SessionLocal() as session:
        service = ManualOutreachService(session, ManualEmailSendRecordRepository(session))
        service.record_external_manual_send_batch(
            ConfirmedExternalManualEmailSendBatchCommand(
                project_id=ids["project"], email_draft_ids=(second,), confirmed=True
            )
        )
        session.commit()
    result = _invoke(ids, tmp_path / "conflict.xlsx")
    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["status"] == "ALREADY_MANUALLY_SENT"
    assert payload["conflicting_draft_id"] == second
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 1
        draft = session.get(EmailDraft, first)
        assert draft is not None and draft.delivery_mode is None


def test_delivery_attempt_rejects_entire_batch(tmp_path: Path) -> None:
    ids = _batch(2)
    first, second = ids["drafts"]
    with SessionLocal() as session:
        draft = session.get(EmailDraft, second)
        assert draft is not None
        session.add(
            EmailDeliveryAttempt(
                email_draft_id=second,
                attempt_key="d" * 64,
                recipient_email=draft.recipient_email,
                envelope_from="sender@example.test",
                header_from_email="sender@example.test",
                message_id="<batch@example.test>",
                content_hash=draft.content_hash,
                transport_name="fake",
                security_mode="STARTTLS",
            )
        )
        session.commit()
    result = _invoke(ids, tmp_path / "attempt.xlsx")
    assert result.exit_code == 1
    payload = json.loads(result.stdout)
    assert payload["status"] == "DELIVERY_ATTEMPT_EXISTS"
    assert payload["conflicting_draft_id"] == second
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0
        draft = session.get(EmailDraft, first)
        assert draft is not None and draft.delivery_mode is None


def test_duplicate_ids_and_wrong_project_are_rejected(tmp_path: Path) -> None:
    ids = _batch(1)
    draft_id = ids["drafts"][0]
    duplicate = dict(ids, drafts=(draft_id, draft_id))
    duplicate_result = _invoke(duplicate, tmp_path / "duplicate.xlsx")
    assert duplicate_result.exit_code == 1
    assert json.loads(duplicate_result.stdout)["status"] == "INVALID_DRAFT"
    wrong = dict(ids, project=int(ids["project"]) + 1000)
    wrong_result = _invoke(wrong, tmp_path / "wrong.xlsx")
    assert wrong_result.exit_code == 1
    assert json.loads(wrong_result.stdout)["status"] == "INVALID_DRAFT"
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0


def test_late_persistence_failure_rolls_back_whole_batch() -> None:
    ids = _batch(2)

    class FailingRepository(ManualEmailSendRecordRepository):
        calls = 0

        def create(self, data):
            record = super().create(data)
            self.calls += 1
            if self.calls == 2:
                raise SQLAlchemyError("late failure")
            return record

    with SessionLocal() as session, pytest.raises(ManualOutreachPersistenceError):
        service = ManualOutreachService(session, FailingRepository(session))
        service.record_external_manual_send_batch(
            ConfirmedExternalManualEmailSendBatchCommand(
                project_id=ids["project"],
                email_draft_ids=ids["drafts"],
                confirmed=True,
            )
        )
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 0
        assert all(
            session.get(EmailDraft, draft_id).delivery_mode is None for draft_id in ids["drafts"]
        )


def test_final_export_failure_reports_recorded_batch(tmp_path: Path) -> None:
    ids = _batch(1)

    class FailingFinalExporter(CRMExcelExportService):
        calls = 0

        def export_drafts(self, *args, **kwargs):
            self.calls += 1
            if self.calls == 2:
                raise CRMExcelExportError("final failure")
            return super().export_drafts(*args, **kwargs)

    result = OutreachBatchWorkflow(
        SessionLocal, exporter=FailingFinalExporter(), clock=lambda: NOW
    ).execute(
        project_id=ids["project"],
        email_draft_ids=ids["drafts"],
        output_file=tmp_path / "failed.xlsx",
        confirmed=True,
    )
    assert result.status is OutreachBatchStatus.MANUAL_SEND_RECORDED_EXPORT_FAILED
    assert [item.email_draft_id for item in result.items] == list(ids["drafts"])
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 1


def test_batch_never_initializes_smtp_or_network(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ids = _batch(2)

    def blocked(*_args: object, **_kwargs: object) -> None:
        raise AssertionError("Batch attempted network delivery.")

    monkeypatch.setattr(smtplib, "SMTP", blocked)
    monkeypatch.setattr(smtplib, "SMTP_SSL", blocked)
    monkeypatch.setattr(socket, "create_connection", blocked)
    result = _invoke(ids, tmp_path / "zero-network.xlsx")
    assert result.exit_code == 0, result.output


def test_rerun_after_success_is_conflict_without_duplicate(tmp_path: Path) -> None:
    ids = _batch(1)
    assert _invoke(ids, tmp_path / "first.xlsx").exit_code == 0
    rerun = _invoke(ids, tmp_path / "second.xlsx")
    assert rerun.exit_code == 1
    assert json.loads(rerun.stdout)["status"] == "ALREADY_MANUALLY_SENT"
    with SessionLocal() as session:
        assert session.scalar(select(func.count()).select_from(ManualEmailSendRecord)) == 1

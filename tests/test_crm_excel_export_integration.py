import os
import subprocess
import sys
from collections.abc import Iterator
from contextlib import contextmanager
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import event, select
from sqlalchemy.engine import Engine
from typer.testing import CliRunner

from app.cli.main import app
from app.core.database.session import SessionLocal
from app.modules.company.models import Company
from app.modules.company_enrichment.models import CompanyEnrichment
from app.modules.contact.models import Contact
from app.modules.email_delivery.manual_models import ManualEmailSendRecord
from app.modules.email_delivery.models import EmailDeliveryAttempt
from app.modules.email_draft.models import EmailDraft
from app.modules.lead.models import Lead
from app.modules.project.models import Project
from app.modules.task.models import Task

runner = CliRunner()
_SNAPSHOT_MODELS = (
    Project,
    Company,
    CompanyEnrichment,
    Contact,
    Lead,
    Task,
    EmailDraft,
    ManualEmailSendRecord,
    EmailDeliveryAttempt,
)


def _snapshot() -> dict[str, tuple[tuple[tuple[str, object], ...], ...]]:
    with SessionLocal() as session:
        return {
            model.__tablename__: tuple(
                tuple(
                    (column.name, getattr(record, column.name))
                    for column in model.__table__.columns
                )
                for record in session.scalars(select(model).order_by(model.id)).all()
            )
            for model in _SNAPSHOT_MODELS
        }


@contextmanager
def _reject_dml(engine: Engine) -> Iterator[None]:
    def before_execute(
        _connection: object,
        _cursor: object,
        statement: str,
        _parameters: object,
        _context: object,
        _executemany: bool,
    ) -> None:
        if statement.lstrip().split(None, 1)[0].upper() in {
            "ALTER",
            "CREATE",
            "DELETE",
            "DROP",
            "INSERT",
            "REPLACE",
            "UPDATE",
        }:
            raise AssertionError("Excel export attempted a database mutation.")

    event.listen(engine, "before_cursor_execute", before_execute)
    try:
        yield
    finally:
        event.remove(engine, "before_cursor_execute", before_execute)


def test_real_cli_export_is_read_only_and_preserves_hillary_semantics(tmp_path: Path) -> None:
    from app.core.database.engine import engine

    now = datetime(2026, 8, 12, tzinfo=UTC)
    with SessionLocal() as session:
        project = Project(name="CRM Export")
        session.add(project)
        session.flush()
        company = Company(
            project_id=project.id,
            name="Simon Wallace Design",
            website="https://simonwallacedesign.com",
        )
        session.add(company)
        session.flush()
        session.add(
            CompanyEnrichment(
                company_id=company.id,
                enrichment_status="SUCCEEDED",
                email="info@simonwallacedesign.com",
                instagram_url="https://instagram.com/simonwallacedesign",
                linkedin_url="https://linkedin.com/company/simon-wallace-design",
                source_url="https://simonwallacedesign.com",
            )
        )
        contact = Contact(
            company_id=company.id,
            first_name="Hillary",
            last_name="Wallace",
            job_title="Principal Designer",
            email="hillary@example.com",
            source="MANUAL_VERIFIED",
        )
        session.add(contact)
        session.flush()
        lead = Lead(company_id=company.id, contact_id=contact.id, status="NEW")
        session.add(lead)
        session.flush()
        historical = Task(lead_id=lead.id, title="Initial outreach", status="DONE")
        current = Task(lead_id=lead.id, title="Follow up", status="TODO")
        session.add_all((historical, current))
        session.flush()
        draft = EmailDraft(
            project_id=project.id,
            company_id=company.id,
            contact_id=contact.id,
            lead_id=lead.id,
            task_id=historical.id,
            recipient_email="hillary@example.com",
            recipient_name="Hillary Wallace",
            recipient_role="Principal Designer",
            sender_name="Operator",
            sender_company="CRM",
            generation_tone="professional",
            generation_purpose="Outreach",
            generation_value_proposition=None,
            subject="Persisted subject",
            text_body="Persisted approved body",
            language="en",
            prompt_version="test",
            provider="fake",
            model="fake",
            context_fingerprint="a" * 64,
            request_fingerprint="b" * 64,
            content_hash="c" * 64,
            status="APPROVED",
            delivery_mode="MANUAL",
            generated_at=now,
            reviewed_at=now,
            approved_at=now,
        )
        session.add(draft)
        session.flush()
        session.add(
            ManualEmailSendRecord(
                project_id=project.id,
                company_id=company.id,
                contact_id=contact.id,
                email_draft_id=draft.id,
                recipient_email="hillary@example.com",
                sent_at=now,
            )
        )
        session.commit()
        project_id = project.id
    before = _snapshot()
    destination = tmp_path / "exports" / "crm.xlsx"
    with _reject_dml(engine):
        result = runner.invoke(
            app,
            [
                "crm",
                "export-excel",
                "--project-id",
                str(project_id),
                "--output-file",
                str(destination),
            ],
        )
    assert result.exit_code == 0, result.output
    assert _snapshot() == before
    workbook = load_workbook(destination)
    headers = {cell.value: cell.column for cell in workbook["Sales Leads"][1]}
    row = workbook["Sales Leads"][2]
    assert row[headers["Current Task ID"] - 1].value == current.id
    assert row[headers["Draft ID"] - 1].value == draft.id
    assert row[headers["Outreach Status"] - 1].value == "MANUALLY_SENT"
    assert row[headers["Company Email"] - 1].value == "info@simonwallacedesign.com"
    assert row[headers["Decision Maker Email"] - 1].value == "hillary@example.com"
    assert row[headers["Recipient Type"] - 1].value == "PERSON"
    assert row[headers["Recipient Email"] - 1].value == "hillary@example.com"
    assert row[headers["Outreach Readiness"] - 1].value == "SENT"
    assert row[headers["Email Source"] - 1].value == "MANUAL_VERIFIED"
    assert row[headers["Due At"] - 1].value is None
    assert row[headers["Email Subject"] - 1].value == "Persisted subject"
    assert row[headers["Email Text"] - 1].value == "Persisted approved body"
    outreach_headers = {cell.value: cell.column for cell in workbook["Outreach"][1]}
    outreach_row = workbook["Outreach"][2]
    assert outreach_row[outreach_headers["Recipient Type"] - 1].value == "PERSON"
    assert outreach_row[outreach_headers["Recipient Email"] - 1].value == "hillary@example.com"
    assert outreach_row[outreach_headers["Contact"] - 1].value == "Hillary Wallace"
    assert workbook["Tasks"].max_row == 3


def test_export_help_is_fresh_process_import_safe() -> None:
    proof = r"""
import smtplib, socket, pydantic_settings, sqlalchemy, sqlalchemy.orm
from typer.testing import CliRunner
counts = {name: 0 for name in (
    'settings', 'engine', 'sessionmaker', 'session', 'smtp', 'network', 'openpyxl'
)}
def blocked(name):
    def fail(*args, **kwargs):
        counts[name] += 1
        raise AssertionError(name)
    return fail
pydantic_settings.BaseSettings.__init__ = blocked('settings')
sqlalchemy.create_engine = blocked('engine')
sqlalchemy.orm.sessionmaker = blocked('sessionmaker')
sqlalchemy.orm.Session.__init__ = blocked('session')
smtplib.SMTP = blocked('smtp'); smtplib.SMTP_SSL = blocked('smtp')
socket.create_connection = blocked('network'); socket.getaddrinfo = blocked('network')
import sys
sys.modules['openpyxl'] = type(
    'Blocked', (), {'__getattr__': lambda self, name: blocked('openpyxl')()}
)()
from app.cli.main import app
results = [
    CliRunner().invoke(app, args).exit_code
    for args in (['--help'], ['crm', '--help'], ['crm', 'export-excel', '--help'])
]
print(results); print(counts)
"""
    environment = os.environ.copy()
    environment["DEBUG"] = "false"
    result = subprocess.run(
        [sys.executable, "-c", proof],
        env=environment,
        check=False,
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, result.stderr
    assert result.stdout.splitlines() == [
        "[0, 0, 0]",
        "{'settings': 0, 'engine': 0, 'sessionmaker': 0, 'session': 0, "
        "'smtp': 0, 'network': 0, 'openpyxl': 0}",
    ]

from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

import app.cli.crm as crm_cli
from app.cli.main import app
from app.modules.crm import CRMOverviewRow
from app.modules.crm.excel_export import (
    CRMExcelDataset,
    CRMExcelExportError,
    CRMExcelExportResult,
    CRMExcelExportService,
    ExcelCompany,
    ExcelContact,
    ExcelOutreach,
    ExcelTask,
)

runner = CliRunner()
NOW = datetime(2026, 8, 12, 2, 32, tzinfo=UTC)


def dataset(*, company_name: str = "Acme | Design | NY") -> CRMExcelDataset:
    row = CRMOverviewRow(
        company_id=11,
        company=company_name,
        contact_id=2,
        contact="[bold]Literal[/bold]",
        role="Principal Designer",
        email="hillary@example.com",
        lead_id=2,
        lead_status="NEW",
        task_id=3,
        task='=HYPERLINK("https://evil.example","click")',
        task_status="TODO",
        draft_id=1,
        draft_task_id=2,
        draft_status="APPROVED",
        outreach_status="MANUALLY_SENT",
        last_sent_at=NOW,
    )
    company = ExcelCompany(
        11,
        1,
        company_name,
        "@SUM(A1:A2)",
        None,
        None,
        None,
        "NEW",
        None,
        "info@example.com",
        "+123456789",
        "https://instagram.com/acme",
        "https://linkedin.com/company/acme",
        "https://acme.test/contact",
        "https://acme.test/about",
        "SUCCEEDED",
        "https://acme.test",
        NOW,
        None,
    )
    contact = ExcelContact(
        2,
        11,
        company_name,
        "+1+1",
        None,
        "Founder",
        "hillary@example.com",
        None,
        None,
        None,
        None,
        "MANUAL_VERIFIED",
        "NEW",
        None,
    )
    outreach = ExcelOutreach(
        1,
        1,
        company_name,
        "+1+1",
        "hillary@example.com",
        "=Subject",
        "APPROVED",
        2,
        "MANUAL",
        "MANUALLY_SENT",
        NOW,
        None,
        None,
        NOW,
        NOW,
        NOW,
        "[link=https://example.com]Company[/link]",
    )
    service = CRMExcelExportService()
    sales = service._build_sales_leads((company,), (contact,), (row,), (outreach,))
    return CRMExcelDataset(
        sales_leads=sales,
        leads=(row,),
        lead_project_ids={11: 1},
        task_due_dates={3: NOW},
        companies=(company,),
        contacts=(contact,),
        tasks=(
            ExcelTask(2, 2, company_name, "+1+1", "Historical", None, "DONE", None),
            ExcelTask(3, 2, company_name, "+1+1", row.task or "", "@SUM(A1:A2)", "TODO", NOW),
        ),
        outreach=(outreach,),
    )


def test_command_registration_help_and_required_output_file() -> None:
    help_result = runner.invoke(app, ["crm", "export-excel", "--help"])
    missing_result = runner.invoke(app, ["crm", "export-excel"])
    assert help_result.exit_code == 0
    assert "--output-file" in help_result.output
    assert "--project-id" in help_result.output
    assert "--company-id" in help_result.output
    assert "--overwrite" in help_result.output
    assert missing_result.exit_code != 0


def test_cli_forwards_filters_overwrite_and_reports_counts(monkeypatch, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    calls: list[tuple[int | None, int | None, Path, bool]] = []
    output = tmp_path / "crm.xlsx"

    def export(
        project: int | None,
        company: int | None,
        path: Path,
        *,
        overwrite: bool,
    ) -> CRMExcelExportResult:
        calls.append((project, company, path, overwrite))
        return CRMExcelExportResult(
            path,
            {name: 0 for name in ("Sales Leads", "Companies", "Contacts", "Tasks", "Outreach")},
        )

    monkeypatch.setattr(crm_cli, "_export_excel", export)
    result = runner.invoke(
        app,
        [
            "crm",
            "export-excel",
            "--project-id",
            "1",
            "--company-id",
            "11",
            "--output-file",
            str(output),
            "--overwrite",
        ],
    )
    assert result.exit_code == 0
    assert calls == [(1, 11, output, True)]
    assert "CRM Excel export created:" in result.output
    assert "Leads: 0" in result.output


def test_cli_reports_safe_known_and_unknown_errors(monkeypatch, tmp_path: Path) -> None:  # type: ignore[no-untyped-def]
    def fail_known(*_args: object, **_kwargs: object) -> CRMExcelExportResult:
        raise CRMExcelExportError("Destination already exists: export.xlsx")

    monkeypatch.setattr(crm_cli, "_export_excel", fail_known)
    known = runner.invoke(app, ["crm", "export-excel", "--output-file", str(tmp_path / "x")])

    def fail_unknown(*_args: object, **_kwargs: object) -> CRMExcelExportResult:
        raise RuntimeError("secret")

    monkeypatch.setattr(crm_cli, "_export_excel", fail_unknown)
    unknown = runner.invoke(app, ["crm", "export-excel", "--output-file", str(tmp_path / "y")])
    assert known.exit_code == 1 and "Destination already exists" in known.stderr
    assert unknown.exit_code == 1 and unknown.stderr == "CRM Excel export failed.\n"
    assert "secret" not in unknown.output


def test_workbook_structure_semantics_and_literal_cells(tmp_path: Path) -> None:
    destination = tmp_path / "nested" / "crm"
    result = CRMExcelExportService()._export_dataset(dataset(), destination)
    assert result.output_file == Path(f"{destination}.xlsx")
    workbook = load_workbook(result.output_file, data_only=False)
    assert workbook.sheetnames == ["Sales Leads", "Companies", "Contacts", "Tasks", "Outreach"]
    leads = workbook["Sales Leads"]
    assert leads.freeze_panes == "A2"
    assert leads.auto_filter.ref == "A1:AK2"
    assert "SalesLeadsTable" in leads.tables
    values = {
        leads.cell(1, column).value: leads.cell(2, column)
        for column in range(1, leads.max_column + 1)
    }
    assert values["Current Task ID"].value == 3
    assert values["Draft ID"].value == 1
    assert values["Recommended Recipient Type"].value == "DECISION_MAKER"
    assert values["Recommended Recipient"].value == "hillary@example.com"
    assert values["Email Subject"].value == "=Subject"
    assert values["Email Subject"].data_type == "s"
    assert values["Last Sent At"].number_format == "yyyy-mm-dd hh:mm:ss"
    assert values["Company Name"].value == "Acme | Design | NY"
    assert values["Decision Maker Name"].value == "+1+1"
    assert values["Current Task"].data_type == "s"
    assert values["Current Task"].value.startswith("=HYPERLINK")
    assert workbook["Companies"]["D2"].data_type == "s"
    assert workbook["Companies"]["D2"].value == "@SUM(A1:A2)"
    assert workbook["Contacts"]["D2"].data_type == "s"
    assert workbook["Tasks"]["F3"].data_type == "s"
    assert workbook["Outreach"]["F2"].data_type == "s"
    assert workbook["Outreach"]["G2"].value == "[link=https://example.com]Company[/link]"
    assert workbook["Outreach"].column_dimensions["G"].width <= 80


def test_empty_workbook_has_headers_without_invalid_tables(tmp_path: Path) -> None:
    empty = CRMExcelDataset((), (), {}, {}, (), (), (), ())
    destination = tmp_path / "empty.xlsx"
    result = CRMExcelExportService()._export_dataset(empty, destination)
    workbook = load_workbook(result.output_file)
    assert result.counts == {name: 0 for name in workbook.sheetnames}
    assert all(sheet.max_row == 1 and not sheet.tables for sheet in workbook.worksheets)


def test_existing_destination_requires_overwrite_and_preserves_good_file(tmp_path: Path) -> None:
    destination = tmp_path / "crm.xlsx"
    destination.write_bytes(b"existing-good-file")
    service = CRMExcelExportService()
    with pytest.raises(CRMExcelExportError, match="Destination already exists"):
        service._export_dataset(dataset(), destination)
    assert destination.read_bytes() == b"existing-good-file"
    service._export_dataset(dataset(), destination, overwrite=True)
    assert load_workbook(destination).sheetnames == list(_SHEET_NAMES_FOR_TEST)


_SHEET_NAMES_FOR_TEST = ("Sales Leads", "Companies", "Contacts", "Tasks", "Outreach")


def test_company_only_and_recipient_priority_rules() -> None:
    service = CRMExcelExportService()
    company = ExcelCompany(
        1,
        1,
        "Company Only",
        None,
        None,
        None,
        None,
        "NEW",
        None,
        "hello@example.com",
        None,
        "https://instagram.com/company",
        None,
        None,
        None,
        "PARTIAL",
        None,
        None,
        None,
    )
    company_only = service._build_sales_leads((company,), (), (), ())[0]
    assert company_only.recommended_recipient_type == "COMPANY"
    assert company_only.recommended_recipient == "hello@example.com"
    assert company_only.decision_maker_name is None
    assert company_only.email_subject is None and company_only.email_text is None

    project_manager = ExcelContact(
        1,
        1,
        company.name,
        "Pat",
        "Manager",
        "Project Manager",
        "pm@example.com",
        None,
        None,
        None,
        None,
        "MANUAL",
        "NEW",
        None,
    )
    founder = ExcelContact(
        2,
        1,
        company.name,
        "Fran",
        "Founder",
        "Founder",
        "founder@example.com",
        None,
        None,
        None,
        None,
        "VERIFIED",
        "NEW",
        None,
    )
    selected = service._build_sales_leads((company,), (project_manager, founder), (), ())[0]
    assert selected.decision_maker_contact_id == founder.id
    assert selected.recommended_recipient_type == "DECISION_MAKER"
    assert selected.recommended_recipient == "founder@example.com"
    assert selected.company_email == "hello@example.com"


def test_person_without_email_falls_back_to_company_and_no_email_remains_visible() -> None:
    service = CRMExcelExportService()
    company = ExcelCompany(
        1,
        1,
        "Fallback",
        None,
        None,
        None,
        None,
        "NEW",
        None,
        "info@example.com",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    principal = ExcelContact(
        1,
        1,
        company.name,
        "Priya",
        None,
        "Principal",
        None,
        None,
        None,
        None,
        None,
        "MANUAL",
        "NEW",
        None,
    )
    selected = service._build_sales_leads((company,), (principal,), (), ())[0]
    assert selected.decision_maker_name == "Priya"
    assert selected.decision_maker_email is None
    assert selected.recommended_recipient_type == "COMPANY"
    assert selected.recommended_recipient == "info@example.com"

    no_email_company = ExcelCompany(
        2,
        1,
        "Social Only",
        None,
        None,
        None,
        None,
        "NEW",
        None,
        None,
        None,
        "https://instagram.com/social",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    )
    no_email = service._build_sales_leads((no_email_company,), (), (), ())[0]
    assert no_email.recommended_recipient_type == "NO_EMAIL"
    assert no_email.recommended_recipient is None

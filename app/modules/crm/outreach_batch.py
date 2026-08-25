from __future__ import annotations

import tempfile
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime
from enum import StrEnum
from pathlib import Path
from typing import Protocol

from sqlalchemy.orm import Session

from app.modules.email_delivery.manual_repository import ManualEmailSendRecordRepository
from app.modules.email_delivery.manual_schemas import (
    ConfirmedExternalManualEmailSendBatchCommand,
    ManualEmailCopyPackage,
    ManualOutreachBatchReason,
)
from app.modules.email_delivery.manual_service import (
    ManualOutreachBatchValidationError,
    ManualOutreachError,
    ManualOutreachService,
)

from .excel_export import CRMExcelExportError, CRMExcelExportService


class SessionFactory(Protocol):
    def __call__(self) -> Session: ...


class OutreachBatchStatus(StrEnum):
    COMPLETE = "COMPLETE"
    CONFIRMATION_REQUIRED = "CONFIRMATION_REQUIRED"
    BATCH_CONFLICT = "BATCH_CONFLICT"
    INVALID_DRAFT = "INVALID_DRAFT"
    ALREADY_MANUALLY_SENT = "ALREADY_MANUALLY_SENT"
    DELIVERY_ATTEMPT_EXISTS = "DELIVERY_ATTEMPT_EXISTS"
    EXPORT_FAILED_BEFORE_MUTATION = "EXPORT_FAILED_BEFORE_MUTATION"
    MANUAL_SEND_RECORDED_EXPORT_FAILED = "MANUAL_SEND_RECORDED_EXPORT_FAILED"
    PERSISTENCE_FAILURE = "PERSISTENCE_FAILURE"


@dataclass(frozen=True, slots=True)
class OutreachBatchItem:
    email_draft_id: int
    company_id: int
    company_name: str
    recipient_email: str
    recipient_type: str
    outreach_status: str
    manual_send_record_id: int | None
    sent_at: datetime | None


@dataclass(frozen=True, slots=True)
class OutreachBatchResult:
    status: OutreachBatchStatus
    project_id: int
    email_draft_ids: tuple[int, ...]
    output_file: Path
    items: tuple[OutreachBatchItem, ...] = ()
    conflicting_draft_id: int | None = None
    message: str | None = None

    def as_dict(self) -> dict[str, object]:
        return {
            "status": self.status.value,
            "project_id": self.project_id,
            "email_draft_ids": list(self.email_draft_ids),
            "output_file": str(self.output_file),
            "conflicting_draft_id": self.conflicting_draft_id,
            "message": self.message,
            "items": [
                {
                    "email_draft_id": item.email_draft_id,
                    "company_id": item.company_id,
                    "company_name": item.company_name,
                    "recipient_email": item.recipient_email,
                    "recipient_type": item.recipient_type,
                    "outreach_status": item.outreach_status,
                    "manual_send_record_id": item.manual_send_record_id,
                    "sent_at": item.sent_at.isoformat() if item.sent_at else None,
                }
                for item in self.items
            ],
        }


class OutreachBatchWorkflow:
    def __init__(
        self,
        session_factory: SessionFactory,
        *,
        exporter: CRMExcelExportService | None = None,
        clock: Callable[[], datetime] | None = None,
    ) -> None:
        self.session_factory = session_factory
        self.exporter = exporter or CRMExcelExportService()
        self.clock = clock

    def execute(
        self,
        *,
        project_id: int,
        email_draft_ids: tuple[int, ...],
        output_file: Path,
        confirmed: bool,
    ) -> OutreachBatchResult:
        destination = self.exporter.normalize_output_file(output_file)
        command = ConfirmedExternalManualEmailSendBatchCommand(
            project_id=project_id,
            email_draft_ids=email_draft_ids,
            confirmed=confirmed,
        )
        preview = self._preview(command, destination)
        if isinstance(preview, OutreachBatchResult):
            return preview
        if not confirmed:
            return self._result(
                OutreachBatchStatus.CONFIRMATION_REQUIRED,
                command,
                destination,
                preview,
                message="Use --confirm to record the entire batch as manually sent.",
            )
        if destination.exists():
            return self._result(
                OutreachBatchStatus.EXPORT_FAILED_BEFORE_MUTATION,
                command,
                destination,
                preview,
                message=f"Destination already exists: {destination}",
            )
        try:
            self._preflight_export(command, destination, preview)
        except Exception:
            return self._result(
                OutreachBatchStatus.EXPORT_FAILED_BEFORE_MUTATION,
                command,
                destination,
                preview,
                message="Outreach batch export preflight failed.",
            )

        recorded = self._record(command, destination)
        if isinstance(recorded, OutreachBatchResult):
            return recorded
        try:
            with self.session_factory() as session:
                self.exporter.export_drafts(
                    session,
                    project_id=project_id,
                    email_draft_ids=email_draft_ids,
                    output_file=destination,
                    overwrite=False,
                )
            self._verify_workbook(destination, recorded, sent=True)
        except Exception:
            return self._result(
                OutreachBatchStatus.MANUAL_SEND_RECORDED_EXPORT_FAILED,
                command,
                destination,
                recorded,
                message="Manual sends were recorded, but the final export failed.",
            )
        return self._result(
            OutreachBatchStatus.COMPLETE,
            command,
            destination,
            recorded,
        )

    def _preflight_export(
        self,
        command: ConfirmedExternalManualEmailSendBatchCommand,
        destination: Path,
        packages: tuple[ManualEmailCopyPackage, ...],
    ) -> None:
        preflight: Path | None = None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                dir=destination.parent,
                prefix=f".{destination.name}.preflight.",
                suffix=".xlsx",
                delete=False,
            ) as handle:
                preflight = Path(handle.name)
            preflight.unlink()
            with self.session_factory() as session:
                self.exporter.export_drafts(
                    session,
                    project_id=command.project_id,
                    email_draft_ids=command.email_draft_ids,
                    output_file=preflight,
                    overwrite=False,
                )
            self._verify_workbook(preflight, packages, sent=False)
        finally:
            if preflight is not None:
                preflight.unlink(missing_ok=True)

    def _preview(
        self,
        command: ConfirmedExternalManualEmailSendBatchCommand,
        destination: Path,
    ) -> tuple[ManualEmailCopyPackage, ...] | OutreachBatchResult:
        try:
            with self.session_factory() as session:
                return ManualOutreachService(
                    session,
                    ManualEmailSendRecordRepository(session),
                    clock=self.clock,
                ).preview_external_manual_send_batch(command)
        except ManualOutreachBatchValidationError as error:
            return self._error_result(error, command, destination)
        except ManualOutreachError:
            return self._result(
                OutreachBatchStatus.INVALID_DRAFT,
                command,
                destination,
                message="Outreach batch validation failed.",
            )

    def _record(
        self,
        command: ConfirmedExternalManualEmailSendBatchCommand,
        destination: Path,
    ) -> tuple[ManualEmailCopyPackage, ...] | OutreachBatchResult:
        session = self.session_factory()
        try:
            packages = ManualOutreachService(
                session,
                ManualEmailSendRecordRepository(session),
                clock=self.clock,
            ).record_external_manual_send_batch(command)
            session.commit()
            return packages
        except ManualOutreachBatchValidationError as error:
            session.rollback()
            return self._error_result(error, command, destination)
        except Exception:
            session.rollback()
            return self._result(
                OutreachBatchStatus.PERSISTENCE_FAILURE,
                command,
                destination,
                message="Outreach batch could not be persisted.",
            )
        finally:
            session.close()

    @staticmethod
    def _error_result(
        error: ManualOutreachBatchValidationError,
        command: ConfirmedExternalManualEmailSendBatchCommand,
        destination: Path,
    ) -> OutreachBatchResult:
        status = {
            ManualOutreachBatchReason.INVALID_DRAFT: OutreachBatchStatus.INVALID_DRAFT,
            ManualOutreachBatchReason.ALREADY_MANUALLY_SENT: (
                OutreachBatchStatus.ALREADY_MANUALLY_SENT
            ),
            ManualOutreachBatchReason.DELIVERY_ATTEMPT_EXISTS: (
                OutreachBatchStatus.DELIVERY_ATTEMPT_EXISTS
            ),
        }[error.reason]
        return OutreachBatchResult(
            status,
            command.project_id,
            command.email_draft_ids,
            destination,
            conflicting_draft_id=error.email_draft_id,
            message=str(error),
        )

    @classmethod
    def _result(
        cls,
        status: OutreachBatchStatus,
        command: ConfirmedExternalManualEmailSendBatchCommand,
        destination: Path,
        packages: tuple[ManualEmailCopyPackage, ...] = (),
        *,
        message: str | None = None,
    ) -> OutreachBatchResult:
        return OutreachBatchResult(
            status,
            command.project_id,
            command.email_draft_ids,
            destination,
            tuple(cls._item(package) for package in packages),
            message=message,
        )

    @staticmethod
    def _item(package: ManualEmailCopyPackage) -> OutreachBatchItem:
        return OutreachBatchItem(
            package.email_draft_id,
            package.company_id,
            package.company_name,
            package.recipient_email,
            package.recipient_type.value,
            package.outreach_status.value,
            package.manual_send_record_id,
            package.sent_at,
        )

    @staticmethod
    def _verify_workbook(
        output_file: Path,
        packages: tuple[ManualEmailCopyPackage, ...],
        *,
        sent: bool,
    ) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(output_file, read_only=True, data_only=False)
        try:
            required = {"Sales Leads", "Companies", "Contacts", "Tasks", "Outreach"}
            if set(workbook.sheetnames) != required:
                raise CRMExcelExportError("Outreach batch workbook sheets are invalid.")

            def read_sheet(
                sheet_name: str,
            ) -> tuple[dict[object, int], list[tuple[object, ...]]]:
                sheet = workbook[sheet_name]
                headers: dict[object, int] = {}
                for position, cell in enumerate(sheet[1]):
                    headers[cell.value] = position
                rows: list[tuple[object, ...]] = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    rows.append(tuple(row))
                return headers, rows

            outreach_headers, outreach_rows = read_sheet("Outreach")
            company_headers, company_rows = read_sheet("Companies")
            task_headers, task_rows = read_sheet("Tasks")
            contact_headers, contact_rows = read_sheet("Contacts")
            sales_headers, sales_rows = read_sheet("Sales Leads")
            for package in packages:
                selected_outreach = [
                    row
                    for row in outreach_rows
                    if row[outreach_headers["Draft ID"]] == package.email_draft_id
                ]
                if len(selected_outreach) != 1:
                    raise CRMExcelExportError("Selected Outreach row is missing or duplicated.")
                outreach_row = selected_outreach[0]
                expected_outreach_status = "MANUALLY_SENT" if sent else package.draft_status
                if (
                    outreach_row[outreach_headers["Company"]] != package.company_name
                    or outreach_row[outreach_headers["Recipient Email"]] != package.recipient_email
                    or outreach_row[outreach_headers["Outreach Status"]] != expected_outreach_status
                ):
                    raise CRMExcelExportError("Selected Outreach row is inconsistent.")
                if sent and outreach_row[outreach_headers["Manual Sent At"]] is None:
                    raise CRMExcelExportError("Manual sent timestamp was not exported.")
                if not any(
                    row[company_headers["Company ID"]] == package.company_id for row in company_rows
                ):
                    raise CRMExcelExportError("Selected Company row is missing.")
                if not any(row[task_headers["Task ID"]] == package.task_id for row in task_rows):
                    raise CRMExcelExportError("Selected Task row is missing.")
                if not any(
                    row[sales_headers["Draft ID"]] == package.email_draft_id
                    and row[sales_headers["Company Name"]] == package.company_name
                    for row in sales_rows
                ):
                    raise CRMExcelExportError("Selected Sales Leads row is inconsistent.")
                if package.contact_id is not None and not any(
                    row[contact_headers["Contact ID"]] == package.contact_id for row in contact_rows
                ):
                    raise CRMExcelExportError("Selected Contact row is missing.")
        finally:
            workbook.close()


__all__ = ["OutreachBatchResult", "OutreachBatchStatus", "OutreachBatchWorkflow"]

from __future__ import annotations

import os
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Protocol, cast

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.modules.company.models import Company
from app.modules.company_enrichment.models import CompanyEnrichment
from app.modules.company_enrichment.normalization import normalize_email
from app.modules.contact.models import Contact
from app.modules.email_delivery.manual_models import ManualEmailSendRecord
from app.modules.email_delivery.models import EmailDeliveryAttempt
from app.modules.email_draft.models import EmailDraft
from app.modules.lead.models import Lead
from app.modules.task.models import Task

from .overview import (
    CRMOverviewRepository,
    CRMOverviewRow,
    CRMOverviewService,
    derive_outreach_status,
)

_SHEET_NAMES = ("Leads", "Companies", "Contacts", "Tasks", "Outreach")
_DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


class _Workbook(Protocol):
    def save(self, filename: str | os.PathLike[str]) -> None: ...


class CRMExcelExportError(RuntimeError):
    """A safe operator-facing Excel export failure."""


@dataclass(frozen=True, slots=True)
class ExcelCompany:
    id: int
    project_id: int
    name: str
    website: str | None
    country: str | None
    city: str | None
    industry: str | None
    status: str
    notes: str | None
    company_email: str | None
    company_phone: str | None
    instagram_url: str | None
    linkedin_url: str | None
    contact_page_url: str | None
    about_page_url: str | None
    enrichment_status: str | None
    enrichment_source: str | None
    enrichment_checked_at: datetime | None
    enrichment_error: str | None


@dataclass(frozen=True, slots=True)
class ExcelContact:
    id: int
    company_id: int
    company: str
    first_name: str | None
    last_name: str | None
    job_title: str | None
    email: str | None
    phone: str | None
    linkedin_url: str | None
    country: str | None
    city: str | None
    source: str | None
    status: str
    notes: str | None


@dataclass(frozen=True, slots=True)
class ExcelTask:
    id: int
    lead_id: int
    company: str
    contact: str | None
    title: str
    description: str | None
    status: str
    due_at: datetime | None


@dataclass(frozen=True, slots=True)
class ExcelOutreach:
    draft_id: int
    project_id: int
    company: str
    contact: str
    recipient_email: str
    subject: str
    draft_status: str
    draft_task_id: int
    delivery_mode: str | None
    outreach_status: str
    manual_sent_at: datetime | None
    automatic_outcome: str | None
    automatic_accepted_at: datetime | None
    generated_at: datetime
    approved_at: datetime | None
    reviewed_at: datetime | None
    email_body: str


@dataclass(frozen=True, slots=True)
class ExcelSalesLead:
    project_id: int
    company_id: int
    company_name: str
    industry: str | None
    country: str | None
    city: str | None
    website: str | None
    company_email: str | None
    company_phone: str | None
    instagram_url: str | None
    facebook_url: str | None
    company_linkedin_url: str | None
    contact_page_url: str | None
    about_page_url: str | None
    enrichment_status: str | None
    enrichment_source: str | None
    decision_maker_contact_id: int | None
    decision_maker_name: str | None
    decision_maker_role: str | None
    decision_maker_email: str | None
    decision_maker_phone: str | None
    decision_maker_linkedin_url: str | None
    decision_maker_source: str | None
    lead_id: int | None
    lead_status: str | None
    current_task_id: int | None
    current_task: str | None
    task_status: str | None
    draft_id: int | None
    draft_status: str | None
    outreach_status: str
    last_sent_at: datetime | None
    recommended_recipient_type: str
    recommended_recipient: str | None
    email_subject: str | None
    email_text: str | None
    notes: str | None


@dataclass(frozen=True, slots=True)
class CRMExcelDataset:
    sales_leads: tuple[ExcelSalesLead, ...]
    leads: tuple[CRMOverviewRow, ...]
    lead_project_ids: dict[int, int]
    task_due_dates: dict[int, datetime | None]
    companies: tuple[ExcelCompany, ...]
    contacts: tuple[ExcelContact, ...]
    tasks: tuple[ExcelTask, ...]
    outreach: tuple[ExcelOutreach, ...]


@dataclass(frozen=True, slots=True)
class CRMExcelExportResult:
    output_file: Path
    counts: dict[str, int]


@dataclass(frozen=True, slots=True)
class _SheetData:
    name: str
    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]
    table_name: str


class CRMExcelExportService:
    """Build a human-readable workbook from the authoritative CRM read model."""

    def load(
        self,
        session: Session,
        *,
        project_id: int | None,
        company_id: int | None,
    ) -> CRMExcelDataset:
        snapshot = CRMOverviewRepository(session).load(project_id, company_id)
        leads = CRMOverviewService().build(snapshot)
        company_ids = [company.id for company in snapshot.companies]
        if not company_ids:
            return CRMExcelDataset((), (), {}, {}, (), (), (), ())

        company_models = session.scalars(
            select(Company).where(Company.id.in_(company_ids)).order_by(Company.name, Company.id)
        ).all()
        enrichment_models = session.scalars(
            select(CompanyEnrichment).where(CompanyEnrichment.company_id.in_(company_ids))
        ).all()
        enrichment_by_company = {
            enrichment.company_id: enrichment for enrichment in enrichment_models
        }
        companies = tuple(
            ExcelCompany(
                company.id,
                company.project_id,
                company.name,
                company.website,
                company.country,
                company.city,
                company.industry,
                company.status,
                company.notes,
                enrichment_by_company[company.id].email
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].phone
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].instagram_url
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].linkedin_url
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].contact_page_url
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].about_page_url
                if company.id in enrichment_by_company
                else None,
                str(enrichment_by_company[company.id].enrichment_status)
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].source_url
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].website_checked_at
                if company.id in enrichment_by_company
                else None,
                enrichment_by_company[company.id].last_error
                if company.id in enrichment_by_company
                else None,
            )
            for company in company_models
        )
        company_by_id = {company.id: company for company in companies}

        contact_models = session.scalars(
            select(Contact).where(Contact.company_id.in_(company_ids)).order_by(Contact.id)
        ).all()
        contacts = tuple(
            ExcelContact(
                contact.id,
                contact.company_id,
                company_by_id[contact.company_id].name,
                contact.first_name,
                contact.last_name,
                contact.job_title,
                contact.email,
                contact.phone,
                contact.linkedin_url,
                contact.country,
                contact.city,
                contact.source,
                contact.status,
                contact.notes,
            )
            for contact in contact_models
        )
        contact_by_id = {contact.id: contact for contact in contacts}

        lead_models = session.scalars(
            select(Lead).where(Lead.company_id.in_(company_ids)).order_by(Lead.id)
        ).all()
        lead_ids = [lead.id for lead in lead_models]
        lead_by_id = {lead.id: lead for lead in lead_models}
        task_models = (
            session.scalars(select(Task).where(Task.lead_id.in_(lead_ids)).order_by(Task.id)).all()
            if lead_ids
            else []
        )
        task_rows: list[ExcelTask] = []
        for task in task_models:
            lead = lead_by_id[task.lead_id]
            company_id = lead.company_id
            if company_id is None:
                raise CRMExcelExportError("A Task Lead is missing its Company scope.")
            contact = contact_by_id.get(lead.contact_id) if lead.contact_id is not None else None
            task_rows.append(
                ExcelTask(
                    task.id,
                    task.lead_id,
                    company_by_id[company_id].name,
                    self._contact_name(contact),
                    task.title,
                    task.description,
                    task.status,
                    task.due_at,
                )
            )
        tasks = tuple(task_rows)

        draft_models = (
            session.scalars(
                select(EmailDraft).where(EmailDraft.lead_id.in_(lead_ids)).order_by(EmailDraft.id)
            ).all()
            if lead_ids
            else []
        )
        draft_ids = [draft.id for draft in draft_models]
        manual_models = (
            session.scalars(
                select(ManualEmailSendRecord)
                .where(ManualEmailSendRecord.email_draft_id.in_(draft_ids))
                .order_by(ManualEmailSendRecord.id)
            ).all()
            if draft_ids
            else []
        )
        attempt_models = (
            session.scalars(
                select(EmailDeliveryAttempt)
                .where(EmailDeliveryAttempt.email_draft_id.in_(draft_ids))
                .order_by(EmailDeliveryAttempt.id)
            ).all()
            if draft_ids
            else []
        )
        manual_by_draft = {record.email_draft_id: record for record in manual_models}
        attempt_by_draft = {record.email_draft_id: record for record in attempt_models}
        snapshot_draft_by_id = {draft.id: draft for draft in snapshot.drafts}
        snapshot_manual_by_draft = {
            record.email_draft_id: record for record in snapshot.manual_sends
        }
        snapshot_attempts_by_draft = {
            record.email_draft_id: record for record in snapshot.delivery_attempts
        }
        outreach: list[ExcelOutreach] = []
        for draft in draft_models:
            status, _last_sent_at = derive_outreach_status(
                snapshot_draft_by_id[draft.id],
                snapshot_manual_by_draft.get(draft.id),
                snapshot_attempts_by_draft.get(draft.id),
            )
            manual = manual_by_draft.get(draft.id)
            attempt = attempt_by_draft.get(draft.id)
            outreach.append(
                ExcelOutreach(
                    draft.id,
                    draft.project_id,
                    company_by_id[draft.company_id].name,
                    draft.recipient_name,
                    draft.recipient_email,
                    draft.subject,
                    draft.status,
                    draft.task_id,
                    draft.delivery_mode,
                    status,
                    manual.sent_at if manual else None,
                    attempt.outcome if attempt else None,
                    attempt.accepted_at if attempt else None,
                    draft.generated_at,
                    draft.approved_at,
                    draft.reviewed_at,
                    draft.text_body,
                )
            )

        outreach_records = tuple(outreach)
        return CRMExcelDataset(
            self._build_sales_leads(companies, contacts, leads, outreach_records),
            leads,
            {company.id: company.project_id for company in snapshot.companies},
            {task.id: task.due_at for task in snapshot.tasks},
            companies,
            contacts,
            tasks,
            outreach_records,
        )

    def export(
        self,
        session: Session,
        *,
        project_id: int | None,
        company_id: int | None,
        output_file: Path,
        overwrite: bool,
    ) -> CRMExcelExportResult:
        destination = self.normalize_output_file(output_file)
        if destination.exists() and not overwrite:
            raise CRMExcelExportError(f"Destination already exists: {destination}")
        dataset = self.load(session, project_id=project_id, company_id=company_id)
        return self._export_dataset(dataset, destination, overwrite=overwrite)

    def _export_dataset(
        self,
        dataset: CRMExcelDataset,
        output_file: Path,
        *,
        overwrite: bool = False,
    ) -> CRMExcelExportResult:
        destination = self.normalize_output_file(output_file)
        if destination.exists() and not overwrite:
            raise CRMExcelExportError(f"Destination already exists: {destination}")
        workbook = self.build_workbook(dataset)
        self._save_atomic(workbook, destination, overwrite=overwrite)
        return CRMExcelExportResult(
            destination,
            {
                "Sales Leads": len(dataset.sales_leads),
                "Companies": len(dataset.companies),
                "Contacts": len(dataset.contacts),
                "Tasks": len(dataset.tasks),
                "Outreach": len(dataset.outreach),
            },
        )

    @staticmethod
    def normalize_output_file(output_file: Path) -> Path:
        return (
            output_file if output_file.suffix.casefold() == ".xlsx" else Path(f"{output_file}.xlsx")
        )

    def build_workbook(self, dataset: CRMExcelDataset) -> _Workbook:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo

        workbook = Workbook()
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        for sheet_data in self._sheet_data(dataset):
            sheet = workbook.create_sheet(sheet_data.name)
            sheet.append(sheet_data.headers)
            for row in sheet_data.rows:
                sheet.append([None] * len(row))
                for column, value in enumerate(row, start=1):
                    cell = sheet.cell(sheet.max_row, column)
                    if isinstance(value, str):
                        cell.value = value
                        cell.data_type = "s"
                    elif isinstance(value, datetime):
                        cell.value = value.replace(tzinfo=None) if value.tzinfo else value
                        cell.number_format = _DATETIME_FORMAT
                    else:
                        cell.value = value
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(vertical="top")
            sheet.freeze_panes = "A2"
            last_column = get_column_letter(len(sheet_data.headers))
            sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
            if sheet.max_row > 1:
                table = Table(
                    displayName=sheet_data.table_name,
                    ref=f"A1:{last_column}{sheet.max_row}",
                )
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                sheet.add_table(table)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for index, header in enumerate(sheet_data.headers, start=1):
                values = [
                    header,
                    *(sheet.cell(row, index).value for row in range(2, sheet.max_row + 1)),
                ]
                measured = max(len(str(value)) if value is not None else 0 for value in values) + 2
                sheet.column_dimensions[get_column_letter(index)].width = min(
                    max(measured, self._minimum_width(header)), self._maximum_width(header)
                )
        return cast(_Workbook, workbook)

    @staticmethod
    def _save_atomic(workbook: _Workbook, destination: Path, *, overwrite: bool) -> None:
        temporary: Path | None = None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            with tempfile.NamedTemporaryFile(
                dir=destination.parent,
                prefix=f".{destination.name}.",
                suffix=".tmp",
                delete=False,
            ) as handle:
                temporary = Path(handle.name)
            workbook.save(temporary)
            if overwrite:
                os.replace(temporary, destination)
            else:
                try:
                    os.link(temporary, destination)
                except FileExistsError as error:
                    raise CRMExcelExportError(
                        f"Destination already exists: {destination}"
                    ) from error
                temporary.unlink()
            temporary = None
        except CRMExcelExportError:
            raise
        except OSError as error:
            raise CRMExcelExportError("CRM Excel output could not be written.") from error
        finally:
            if temporary is not None:
                temporary.unlink(missing_ok=True)

    def _sheet_data(self, dataset: CRMExcelDataset) -> tuple[_SheetData, ...]:
        return (
            _SheetData(
                "Sales Leads",
                (
                    "Project ID",
                    "Company ID",
                    "Company Name",
                    "Industry",
                    "Country",
                    "City",
                    "Website",
                    "Company Email",
                    "Company Phone",
                    "Instagram",
                    "Facebook",
                    "Company LinkedIn",
                    "Contact Page",
                    "About Page",
                    "Enrichment Status",
                    "Enrichment Source",
                    "Decision Maker Contact ID",
                    "Decision Maker Name",
                    "Decision Maker Role",
                    "Decision Maker Email",
                    "Decision Maker Phone",
                    "Decision Maker LinkedIn",
                    "Decision Maker Source",
                    "Lead ID",
                    "Lead Status",
                    "Current Task ID",
                    "Current Task",
                    "Task Status",
                    "Draft ID",
                    "Draft Status",
                    "Outreach Status",
                    "Last Sent At",
                    "Recommended Recipient Type",
                    "Recommended Recipient",
                    "Email Subject",
                    "Email Text",
                    "Notes",
                ),
                tuple(
                    (
                        item.project_id,
                        item.company_id,
                        item.company_name,
                        item.industry,
                        item.country,
                        item.city,
                        item.website,
                        item.company_email,
                        item.company_phone,
                        item.instagram_url,
                        item.facebook_url,
                        item.company_linkedin_url,
                        item.contact_page_url,
                        item.about_page_url,
                        item.enrichment_status,
                        item.enrichment_source,
                        item.decision_maker_contact_id,
                        item.decision_maker_name,
                        item.decision_maker_role,
                        item.decision_maker_email,
                        item.decision_maker_phone,
                        item.decision_maker_linkedin_url,
                        item.decision_maker_source,
                        item.lead_id,
                        item.lead_status,
                        item.current_task_id,
                        item.current_task,
                        item.task_status,
                        item.draft_id,
                        item.draft_status,
                        item.outreach_status,
                        item.last_sent_at,
                        item.recommended_recipient_type,
                        item.recommended_recipient,
                        item.email_subject,
                        item.email_text,
                        item.notes,
                    )
                    for item in dataset.sales_leads
                ),
                "SalesLeadsTable",
            ),
            _SheetData(
                "Companies",
                (
                    "Company ID",
                    "Project ID",
                    "Name",
                    "Website",
                    "Industry",
                    "Country",
                    "City",
                    "Status",
                    "Company Email",
                    "Phone",
                    "Instagram",
                    "Facebook",
                    "LinkedIn",
                    "Contact Page",
                    "About Page",
                    "Enrichment Status",
                    "Enrichment Source",
                    "Enrichment Checked At",
                    "Enrichment Error",
                    "Notes",
                ),
                tuple(
                    (
                        item.id,
                        item.project_id,
                        item.name,
                        item.website,
                        item.industry,
                        item.country,
                        item.city,
                        item.status,
                        item.company_email,
                        item.company_phone,
                        item.instagram_url,
                        None,
                        item.linkedin_url,
                        item.contact_page_url,
                        item.about_page_url,
                        item.enrichment_status,
                        item.enrichment_source,
                        item.enrichment_checked_at,
                        item.enrichment_error,
                        item.notes,
                    )
                    for item in dataset.companies
                ),
                "CompaniesTable",
            ),
            _SheetData(
                "Contacts",
                (
                    "Contact ID",
                    "Company ID",
                    "Company",
                    "First Name",
                    "Last Name",
                    "Full Name",
                    "Role",
                    "Email",
                    "Phone",
                    "LinkedIn",
                    "Country",
                    "City",
                    "Source",
                    "Status",
                    "Notes",
                ),
                tuple(
                    (
                        item.id,
                        item.company_id,
                        item.company,
                        item.first_name,
                        item.last_name,
                        self._contact_name(item),
                        item.job_title,
                        item.email,
                        item.phone,
                        item.linkedin_url,
                        item.country,
                        item.city,
                        item.source,
                        item.status,
                        item.notes,
                    )
                    for item in dataset.contacts
                ),
                "ContactsTable",
            ),
            _SheetData(
                "Tasks",
                (
                    "Task ID",
                    "Lead ID",
                    "Company",
                    "Contact",
                    "Title",
                    "Description",
                    "Status",
                    "Due At",
                ),
                tuple(
                    (
                        item.id,
                        item.lead_id,
                        item.company,
                        item.contact,
                        item.title,
                        item.description,
                        item.status,
                        item.due_at,
                    )
                    for item in dataset.tasks
                ),
                "TasksTable",
            ),
            _SheetData(
                "Outreach",
                (
                    "Draft ID",
                    "Project ID",
                    "Company",
                    "Contact",
                    "Recipient Email",
                    "Subject",
                    "Email Body",
                    "Draft Status",
                    "Draft Task ID",
                    "Delivery Mode",
                    "Outreach Status",
                    "Manual Sent At",
                    "Automatic Outcome",
                    "Automatic Accepted At",
                    "Generated At",
                    "Reviewed At",
                    "Approved At",
                ),
                tuple(
                    (
                        item.draft_id,
                        item.project_id,
                        item.company,
                        item.contact,
                        item.recipient_email,
                        item.subject,
                        item.email_body,
                        item.draft_status,
                        item.draft_task_id,
                        item.delivery_mode,
                        item.outreach_status,
                        item.manual_sent_at,
                        item.automatic_outcome,
                        item.automatic_accepted_at,
                        item.generated_at,
                        item.reviewed_at,
                        item.approved_at,
                    )
                    for item in dataset.outreach
                ),
                "OutreachTable",
            ),
        )

    @staticmethod
    def _contact_name(contact: ExcelContact | None) -> str | None:
        if contact is None:
            return None
        return " ".join(part for part in (contact.first_name, contact.last_name) if part) or None

    def _build_sales_leads(
        self,
        companies: tuple[ExcelCompany, ...],
        contacts: tuple[ExcelContact, ...],
        rows: tuple[CRMOverviewRow, ...],
        outreach: tuple[ExcelOutreach, ...],
    ) -> tuple[ExcelSalesLead, ...]:
        contacts_by_company: dict[int, list[ExcelContact]] = {}
        rows_by_company: dict[int, list[CRMOverviewRow]] = {}
        outreach_by_draft = {record.draft_id: record for record in outreach}
        for contact in contacts:
            contacts_by_company.setdefault(contact.company_id, []).append(contact)
        for row in rows:
            rows_by_company.setdefault(row.company_id, []).append(row)
        result: list[ExcelSalesLead] = []
        for company in companies:
            primary = self._select_primary_contact(contacts_by_company.get(company.id, []))
            company_rows = rows_by_company.get(company.id, [])
            selected_row = self._select_master_row(company_rows, primary)
            draft = (
                outreach_by_draft.get(selected_row.draft_id)
                if selected_row and selected_row.draft_id is not None
                else None
            )
            decision_email = (
                primary.email if primary and self._usable_email(primary.email) else None
            )
            company_email = company.company_email
            usable_company_email = company_email if self._usable_email(company_email) else None
            if decision_email is not None:
                recipient_type, recipient = "DECISION_MAKER", decision_email
            elif usable_company_email is not None:
                recipient_type, recipient = "COMPANY", usable_company_email
            else:
                recipient_type, recipient = "NO_EMAIL", None
            result.append(
                ExcelSalesLead(
                    company.project_id,
                    company.id,
                    company.name,
                    company.industry,
                    company.country,
                    company.city,
                    company.website,
                    company.company_email,
                    company.company_phone,
                    company.instagram_url,
                    None,
                    company.linkedin_url,
                    company.contact_page_url,
                    company.about_page_url,
                    company.enrichment_status,
                    company.enrichment_source,
                    primary.id if primary else None,
                    self._contact_name(primary),
                    primary.job_title if primary else None,
                    primary.email if primary else None,
                    primary.phone if primary else None,
                    primary.linkedin_url if primary else None,
                    primary.source if primary else None,
                    selected_row.lead_id if selected_row else None,
                    selected_row.lead_status if selected_row else None,
                    selected_row.task_id if selected_row else None,
                    selected_row.task if selected_row else None,
                    selected_row.task_status if selected_row else None,
                    selected_row.draft_id if selected_row else None,
                    selected_row.draft_status if selected_row else None,
                    selected_row.outreach_status if selected_row else "NO_DRAFT",
                    selected_row.last_sent_at if selected_row else None,
                    recipient_type,
                    recipient,
                    draft.subject if draft else None,
                    draft.email_body if draft else None,
                    company.notes,
                )
            )
        return tuple(result)

    @classmethod
    def _select_primary_contact(cls, contacts: list[ExcelContact]) -> ExcelContact | None:
        ranked = [
            (rank, contact.id, contact)
            for contact in contacts
            if (rank := cls._role_rank(contact.job_title)) is not None
        ]
        if not ranked:
            return None
        return min(ranked, key=lambda item: (item[0], item[1]))[2]

    @staticmethod
    def _select_master_row(
        rows: list[CRMOverviewRow], primary: ExcelContact | None
    ) -> CRMOverviewRow | None:
        if primary is not None:
            matching = [row for row in rows if row.contact_id == primary.id]
            return min(matching, key=lambda row: row.lead_id or 0, default=None)
        unassigned = [row for row in rows if row.contact_id is None]
        return min(unassigned, key=lambda row: row.lead_id or 0, default=None)

    @staticmethod
    def _role_rank(job_title: str | None) -> int | None:
        if not job_title:
            return None
        title = " ".join(job_title.casefold().replace("-", " ").split())
        if any(alias in title for alias in ("co founder", "cofounder")):
            return 1
        priorities = (
            (0, ("founder",)),
            (2, ("owner",)),
            (3, ("managing principal",)),
            (4, ("principal",)),
            (5, ("partner",)),
            (6, ("president",)),
            (7, ("chief executive officer", "ceo")),
            (8, ("creative director",)),
            (9, ("design director",)),
            (10, ("studio director",)),
            (11, ("managing director",)),
            (12, ("procurement", "purchasing")),
            (13, ("senior project manager",)),
            (14, ("project manager",)),
        )
        return next(
            (rank for rank, aliases in priorities if any(alias in title for alias in aliases)),
            None,
        )

    @staticmethod
    def _usable_email(value: str | None) -> bool:
        try:
            return normalize_email(value) is not None
        except ValueError:
            return False

    @staticmethod
    def _minimum_width(header: str) -> int:
        return 10 if header.endswith("ID") else 14

    @staticmethod
    def _maximum_width(header: str) -> int:
        if header.endswith("ID"):
            return 12
        if "Date" in header or " At" in header:
            return 24
        if header in {"Description", "Email Body", "Email Text", "Notes"}:
            return 80
        if header in {"Website", "LinkedIn", "Current Task", "Title", "Subject"}:
            return 60
        if header == "Email" or "Email" in header:
            return 45
        return 40
